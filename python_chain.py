import requests
import json
import pandas as pd
from pandas.io.json import build_table_schema
from glom import glom
import streamlit as st

#import xlwings as xw # Works only on Windows
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Pandas Output Format
pd.set_option('display.width', 1500)
pd.set_option('display.max_rows', 1500)
pd.set_option('display.max_columns', 75)

# Variables
url = 'https://www.nseindia.com/api/option-chain-indices?symbol=BANKNIFTY'
expiry = '30-Sep-2021'
excel_file = 'option_chain_analysis.xlsx'
#wb = xw.Book(excel_file)
#sheet_oi_single = wb.sheets('OIData')
wb = Workbook()
ws1 = wb.active
#wb = load_workbook(excel_file)

headers = {
    'User-Agent' : "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:87.0) Gecko/20100101 Firefox/87.0",
    'Upgrade-Insecure-Requests': "1",
    'Accept': "applicationjson, text/javascript, */*; q=0.01",
    'Accept-Language': "en-GB,en;q=0.5",
    'Accept-Encoding': "gzip, deflate, br",'DNT': "1"
}
cookies = {
    'bm_sv' : '3A8FFBF5409C04DC6D84B39EDFACC362~xQD33G0CZc7uMVaYCLmu6ywyOe8JBy7M3MV5oXanx62m3nKvijYB4NLbD2wxDgkZ8Q5j+JtNILBmVW5FO7j2MLPjj+j0p4Qsudo/EiYKrfUXebMbJDcwOi/j2Nx/VWCuRCj23qBZdXNU7X1rc3mEZU+otX36QVVDDCYhOJR6Ses=',
    'nseappid':	"eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJhcGkubnNlIiwiYXVkIjoiYXBpLm5zZSIsImlhdCI6MTYxOTkzNDMwMSwiZXhwIjoxNjE5OTM3OTAxfQ.rllVot8rOsivraGNC6e5-CA8ECdUzeXA_MHjS9qOKuM",
    'nsit': "DrveQOGzi0Z8hthvx2IDCcf6"
}
session = requests.session()

for cookie in cookies:
    if 'bm_sv' in cookie or 'nseappid' in cookie or 'nsit' in cookie:
        session.cookies.set(cookie, cookies[cookie])


#df = pd.DataFrame().from_dict(data['records']['data'])
#print(df.tail())
#print(data.content)

def fetch_oi():

    expiry = '30-Sep-2021'
    data = session.get(url, headers=headers, timeout=25)
    #print(data.headers)
    print(data.status_code)

    if data.status_code == 1:
        data = data.json()
        #print(type(data))
        with open("oidata.json", "w") as files:
            files.write(json.dumps(data, sort_keys=True))
    else:
        #with open('oidata.json','r') as f:
        #    data = json.loads(f.read())
        data = pd.read_json('oidata.json')
        st.write(data)
        #print(type(data))

    data.to_excel("option_chain_analysis.xlsx", sheet_name='OIData')
    #build_table_schema(data)
    #ce_values = pd.json_normalize(data)
    #ce_value.to_excel("option_chain_analysis.xlsx", sheet_name='CEValues')

"""
    ws2 = wb.create_sheet(title='OIData')
    ws2["A2"].value = data
    wb.save(filename=excel_file)

    if expiry:
        ce_values = [data['CE'] for data in data['records']['data'] if "CE" in data and str(data['expiryDate']).lower() == str(expiry).lower()]
        pe_values = [data['PE'] for data in data['records']['data'] if "PE" in data and str(data['expiryDate']).lower() == str(expiry).lower()]
    else:
        ce_values = [data['CE'] for data in data['filtered']['data'] if "CE" in data]
        pe_values = [data['PE'] for data in data['filtered']['data'] if "PE" in data]

    print(ce_values)

    ce_data = pd.DataFrame(ce_values)
    pe_data = pd.DataFrame(pe_values)

    print(ce_data)
    #print(ce_data[['strikePrice','lastPrice']])
    #print(pe_data[['strikePrice','lastPrice']])"""


def main():
    fetch_oi()

if __name__ == '__main__':
    main()
