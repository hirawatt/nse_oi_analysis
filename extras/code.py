
#import xlwings as xw # Works only on Windows
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# Pandas Output Format
pd.set_option('display.width', 1500)
pd.set_option('display.max_rows', 1500)
pd.set_option('display.max_columns', 75)

#wb = xw.Book(excel_file)
#sheet_oi_single = wb.sheets('OIData')

wb = Workbook()
ws1 = wb.active
wb = load_workbook(excel_file)
ws2 = wb.create_sheet(title='OIData')
ws2["A2"].value = data
wb.save(filename=excel_file)

    if expiry:
        ce_values = [data['CE'] for data in data['records']['data'] if "CE" in data and str(data['expiryDate']).lower() == str(expiry).lower()]
        pe_values = [data['PE'] for data in data['records']['data'] if "PE" in data and str(data['expiryDate']).lower() == str(expiry).lower()]
    else:
        ce_values = [data['CE'] for data in data['filtered']['data'] if "CE" in data]
        pe_values = [data['PE'] for data in data['filtered']['data'] if "PE" in data]

ce_data = pd.DataFrame(ce_values)
pe_data = pd.DataFrame(pe_values)

data.to_excel("option_chain_analysis.xlsx", sheet_name='OIData')
build_table_schema(data)
ce_values = pd.json_normalize(data)
ce_value.to_excel("option_chain_analysis.xlsx", sheet_name='CEValues')

with open('oidata.json','r') as f:
    data = json.loads(f.read())

    #oidata = pd.DataFrame(glom(data, 'records.data'))
    #ce_values = pd.DataFrame(glom(oidata, 'CE'))
    #pe_values = pd.DataFrame(glom(oidata, 'PE'))
